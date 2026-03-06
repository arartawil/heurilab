"""
Styled Excel workbooks: Results.xlsx, Wilcoxon.xlsx, Friedman.xlsx
"""

import os
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from scipy.stats import rankdata
from typing import Dict, List, Tuple

from heurilab.stats.tests import wilcoxon_test, friedman_test, nemenyi_pairwise

# ── Style constants ──────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_PROPOSED_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _style_header(ws, row: int, max_col: int):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER


def _style_cell(ws, row: int, col: int, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if fill:
        cell.fill = fill


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


# ═══════════════════════════════════════════════════════════════════════
#  Results.xlsx
# ═══════════════════════════════════════════════════════════════════════

def generate_results_excel(results_data: Dict[str, Dict[str, Dict[str, float]]],
                           algo_names: List[str],
                           suites: Dict[str, List[str]],
                           output_dir: str):
    """
    results_data[func_name][algo_name] = {'mean': ..., 'std': ..., 'best': ..., 'worst': ..., 'median': ...}
    suites = {'Classical': [func_names], 'CEC2017': [...], ...}
    """
    wb = Workbook()
    wb.remove(wb.active)
    proposed = algo_names[0]
    all_func_names = []

    for category, func_names in suites.items():
        ws = wb.create_sheet(title=category[:31])
        all_func_names.extend(func_names)
        _write_results_sheet(ws, func_names, algo_names, results_data, proposed)

    # "All Functions" sheet
    if len(suites) > 1:
        ws = wb.create_sheet(title="All Functions")
        all_fns = []
        for fns in suites.values():
            all_fns.extend(fns)
        _write_results_sheet(ws, all_fns, algo_names, results_data, proposed)

    # Ranking sheet
    ws_rank = wb.create_sheet(title="Ranking")
    _write_ranking_sheet(ws_rank, all_func_names, algo_names, results_data)

    path = os.path.join(output_dir, "Excel Files")
    os.makedirs(path, exist_ok=True)
    wb.save(os.path.join(path, "Results.xlsx"))


def _write_results_sheet(ws, func_names, algo_names, results_data, proposed):
    # Header
    ws.cell(row=1, column=1, value="Function")
    for j, algo in enumerate(algo_names):
        ws.cell(row=1, column=j + 2, value=f"{algo} (Mean±Std)")
    _style_header(ws, 1, len(algo_names) + 1)

    for i, func in enumerate(func_names):
        row = i + 2
        ws.cell(row=row, column=1, value=func)
        _style_cell(ws, row, 1)

        means = []
        for j, algo in enumerate(algo_names):
            d = results_data.get(func, {}).get(algo, {})
            m = d.get("mean", float("inf"))
            s = d.get("std", 0)
            means.append(m)
            ws.cell(row=row, column=j + 2, value=f"{m:.4e}±{s:.4e}")
            fill = _PROPOSED_FILL if algo == proposed else None
            _style_cell(ws, row, j + 2, fill=fill)

        # Highlight best mean
        if means:
            best_idx = int(np.argmin(means))
            _style_cell(ws, row, best_idx + 2, fill=_GREEN_FILL)

    _auto_width(ws)


def _write_ranking_sheet(ws, func_names, algo_names, results_data):
    n_algos = len(algo_names)
    ws.cell(row=1, column=1, value="Algorithm")
    ws.cell(row=1, column=2, value="Avg Rank")
    ws.cell(row=1, column=3, value="Win Count")
    _style_header(ws, 1, 3)

    # Compute ranks per function
    all_ranks = np.zeros(n_algos)
    win_counts = np.zeros(n_algos, dtype=int)

    for func in func_names:
        means = []
        for algo in algo_names:
            d = results_data.get(func, {}).get(algo, {})
            means.append(d.get("mean", float("inf")))
        ranks = rankdata(means)
        all_ranks += ranks
        best_idx = int(np.argmin(means))
        win_counts[best_idx] += 1

    n_funcs = max(len(func_names), 1)
    avg_ranks = all_ranks / n_funcs

    for j, algo in enumerate(algo_names):
        row = j + 2
        ws.cell(row=row, column=1, value=algo)
        ws.cell(row=row, column=2, value=f"{avg_ranks[j]:.3f}")
        ws.cell(row=row, column=3, value=int(win_counts[j]))
        for c in range(1, 4):
            _style_cell(ws, row, c)

    _auto_width(ws)


# ═══════════════════════════════════════════════════════════════════════
#  Wilcoxon.xlsx
# ═══════════════════════════════════════════════════════════════════════

def generate_wilcoxon_excel(raw_data: Dict[str, Dict[str, List[float]]],
                            algo_names: List[str],
                            suites: Dict[str, List[str]],
                            output_dir: str):
    """
    raw_data[func_name][algo_name] = [fitness_run_1, ..., fitness_run_N]
    """
    wb = Workbook()
    wb.remove(wb.active)
    proposed = algo_names[0]
    competitors = algo_names[1:]

    summary_wins = {c: 0 for c in competitors}
    summary_ties = {c: 0 for c in competitors}
    summary_losses = {c: 0 for c in competitors}

    for category, func_names in suites.items():
        ws = wb.create_sheet(title=category[:31])
        # Header
        ws.cell(row=1, column=1, value="Function")
        col = 2
        for comp in competitors:
            ws.cell(row=1, column=col, value=f"p-value vs {comp}")
            ws.cell(row=1, column=col + 1, value="Winner")
            col += 2
        _style_header(ws, 1, col - 1)

        for i, func in enumerate(func_names):
            row = i + 2
            ws.cell(row=row, column=1, value=func)
            _style_cell(ws, row, 1)

            prop_vals = raw_data.get(func, {}).get(proposed, [])
            c = 2
            for comp in competitors:
                comp_vals = raw_data.get(func, {}).get(comp, [])
                p_val, winner = wilcoxon_test(prop_vals, comp_vals)

                ws.cell(row=row, column=c, value=f"{p_val:.4e}")
                ws.cell(row=row, column=c + 1, value=winner)

                if winner == "proposed":
                    fill = _GREEN_FILL
                    summary_wins[comp] += 1
                elif winner == "competitor":
                    fill = _RED_FILL
                    summary_losses[comp] += 1
                else:
                    fill = _YELLOW_FILL
                    summary_ties[comp] += 1

                _style_cell(ws, row, c, fill=fill)
                _style_cell(ws, row, c + 1, fill=fill)
                c += 2

        _auto_width(ws)

    # Summary sheet
    ws = wb.create_sheet(title="Summary")
    ws.cell(row=1, column=1, value="Competitor")
    ws.cell(row=1, column=2, value="Wins (+)")
    ws.cell(row=1, column=3, value="Ties (=)")
    ws.cell(row=1, column=4, value="Losses (-)")
    ws.cell(row=1, column=5, value="Total (W/T/L)")
    _style_header(ws, 1, 5)

    for i, comp in enumerate(competitors):
        row = i + 2
        w, t, l = summary_wins[comp], summary_ties[comp], summary_losses[comp]
        ws.cell(row=row, column=1, value=comp)
        ws.cell(row=row, column=2, value=w)
        ws.cell(row=row, column=3, value=t)
        ws.cell(row=row, column=4, value=l)
        ws.cell(row=row, column=5, value=f"{w}/{t}/{l}")
        for c in range(1, 6):
            _style_cell(ws, row, c)

    _auto_width(ws)

    path = os.path.join(output_dir, "Excel Files")
    os.makedirs(path, exist_ok=True)
    wb.save(os.path.join(path, "Wilcoxon.xlsx"))


# ═══════════════════════════════════════════════════════════════════════
#  Friedman.xlsx
# ═══════════════════════════════════════════════════════════════════════

def generate_friedman_excel(raw_data: Dict[str, Dict[str, List[float]]],
                            algo_names: List[str],
                            suites: Dict[str, List[str]],
                            output_dir: str):
    """
    Generate Friedman test workbook with mean ranks and Nemenyi post-hoc.
    """
    wb = Workbook()
    wb.remove(wb.active)

    all_funcs = []
    for fns in suites.values():
        all_funcs.extend(fns)

    # ── Sheet 1: Mean Ranks ──
    ws1 = wb.create_sheet(title="Mean Ranks")
    ws1.cell(row=1, column=1, value="Category")
    for j, algo in enumerate(algo_names):
        ws1.cell(row=1, column=j + 2, value=algo)
    ws1.cell(row=1, column=len(algo_names) + 2, value="Chi2")
    ws1.cell(row=1, column=len(algo_names) + 3, value="p-value")
    _style_header(ws1, 1, len(algo_names) + 3)

    row = 2
    for category, func_names in suites.items():
        chi2, p_val, mean_ranks = friedman_test(raw_data, algo_names, func_names)
        ws1.cell(row=row, column=1, value=category)
        for j, algo in enumerate(algo_names):
            ws1.cell(row=row, column=j + 2, value=f"{mean_ranks[algo]:.3f}")
            _style_cell(ws1, row, j + 2)
        ws1.cell(row=row, column=len(algo_names) + 2, value=f"{chi2:.4f}")
        ws1.cell(row=row, column=len(algo_names) + 3, value=f"{p_val:.4e}")
        _style_cell(ws1, row, 1)
        _style_cell(ws1, row, len(algo_names) + 2)
        _style_cell(ws1, row, len(algo_names) + 3)
        row += 1

    # Overall
    chi2, p_val, mean_ranks = friedman_test(raw_data, algo_names, all_funcs)
    ws1.cell(row=row, column=1, value="Overall")
    ws1.cell(row=row, column=1).font = Font(bold=True)
    for j, algo in enumerate(algo_names):
        ws1.cell(row=row, column=j + 2, value=f"{mean_ranks[algo]:.3f}")
        _style_cell(ws1, row, j + 2)
    ws1.cell(row=row, column=len(algo_names) + 2, value=f"{chi2:.4f}")
    ws1.cell(row=row, column=len(algo_names) + 3, value=f"{p_val:.4e}")
    _style_cell(ws1, row, len(algo_names) + 2)
    _style_cell(ws1, row, len(algo_names) + 3)
    _auto_width(ws1)

    # ── Sheet 2: Nemenyi Post-Hoc ──
    ws2 = wb.create_sheet(title="Nemenyi Post-Hoc")
    ws2.cell(row=1, column=1, value="")
    for j, algo in enumerate(algo_names):
        ws2.cell(row=1, column=j + 2, value=algo)
    _style_header(ws2, 1, len(algo_names) + 1)

    pairwise = nemenyi_pairwise(mean_ranks, len(all_funcs), algo_names)

    for i, algo_i in enumerate(algo_names):
        row = i + 2
        ws2.cell(row=row, column=1, value=algo_i)
        _style_cell(ws2, row, 1)
        for j, algo_j in enumerate(algo_names):
            diff, sig = pairwise[(algo_i, algo_j)]
            ws2.cell(row=row, column=j + 2, value=f"{diff:.3f}")
            fill = _RED_FILL if sig else _GREEN_FILL if i != j else None
            _style_cell(ws2, row, j + 2, fill=fill)

    _auto_width(ws2)

    path = os.path.join(output_dir, "Excel Files")
    os.makedirs(path, exist_ok=True)
    wb.save(os.path.join(path, "Friedman.xlsx"))
