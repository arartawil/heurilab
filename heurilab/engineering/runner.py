"""
Runner for engineering design problems.
"""

import os
import time
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Tuple, Type

from heurilab.exporters.plots import plot_convergence, plot_boxplot
from heurilab.exporters.csv_export import _pad_or_trim
from heurilab.engineering.problems import ENGINEERING_PROBLEMS

# ── Style constants ──────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = _THIN_BORDER


def _style_cell(ws, row, col, fill=None):
    cell = ws.cell(row=row, column=col)
    cell.border = _THIN_BORDER
    cell.alignment = Alignment(horizontal="center")
    if fill:
        cell.fill = fill


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def run_engineering_problems(algo_name: str, algo_class: Type,
                             output_dir: str,
                             pop_size: int = 50, max_iter: int = 500,
                             n_runs: int = 30):
    """
    Run engineering design problems using the proposed algorithm only.
    Generates Excel + plots per problem.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for prob_name, obj_func, dim, lb, ub in ENGINEERING_PROBLEMS:
        print(f"  Engineering: {prob_name}")
        lb_arr = np.array(lb)
        ub_arr = np.array(ub)

        best_fitnesses = []
        best_solutions = []
        all_convergences = []

        for run in range(n_runs):
            algo = algo_class(pop_size=pop_size, dim=dim,
                              lb=lb_arr, ub=ub_arr,
                              max_iter=max_iter, obj_func=obj_func)
            sol, fit, conv = algo.optimize()
            conv = list(conv)
            conv = _pad_or_trim(conv, max_iter + 1)
            best_fitnesses.append(fit)
            best_solutions.append(list(sol))
            all_convergences.append(conv)

        arr = np.array(best_fitnesses)
        best_idx = int(np.argmin(arr))

        # ── Summary sheet ──
        ws = wb.create_sheet(title=f"{prob_name[:25]} Summary")
        ws.cell(row=1, column=1, value="Metric")
        ws.cell(row=1, column=2, value="Value")
        _style_header(ws, 1, 2)
        metrics = [("Best", np.min(arr)), ("Mean", np.mean(arr)),
                   ("Worst", np.max(arr)), ("Std", np.std(arr)),
                   ("Median", np.median(arr))]
        for i, (name, val) in enumerate(metrics):
            ws.cell(row=i+2, column=1, value=name)
            ws.cell(row=i+2, column=2, value=f"{val:.10e}")
            _style_cell(ws, i+2, 1)
            _style_cell(ws, i+2, 2)
        _auto_width(ws)

        # ── Optimal Variables sheet ──
        ws2 = wb.create_sheet(title=f"{prob_name[:25]} Vars")
        ws2.cell(row=1, column=1, value="Variable")
        ws2.cell(row=1, column=2, value="Optimal Value")
        _style_header(ws2, 1, 2)
        for v in range(dim):
            ws2.cell(row=v+2, column=1, value=f"x{v+1}")
            ws2.cell(row=v+2, column=2, value=f"{best_solutions[best_idx][v]:.10e}")
            _style_cell(ws2, v+2, 1)
            _style_cell(ws2, v+2, 2)
        _auto_width(ws2)

        # ── All 30 Runs sheet ──
        ws3 = wb.create_sheet(title=f"{prob_name[:25]} Runs")
        ws3.cell(row=1, column=1, value="Run")
        ws3.cell(row=1, column=2, value="Best Fitness")
        for v in range(dim):
            ws3.cell(row=1, column=v+3, value=f"x{v+1}")
        _style_header(ws3, 1, dim + 2)
        for r in range(n_runs):
            ws3.cell(row=r+2, column=1, value=r+1)
            ws3.cell(row=r+2, column=2, value=f"{best_fitnesses[r]:.10e}")
            for v in range(dim):
                ws3.cell(row=r+2, column=v+3, value=f"{best_solutions[r][v]:.10e}")
            for c in range(1, dim + 3):
                _style_cell(ws3, r+2, c)
        _auto_width(ws3)

        # ── Convergence data sheet ──
        ws4 = wb.create_sheet(title=f"{prob_name[:25]} Conv")
        mean_conv = np.mean(all_convergences, axis=0)
        ws4.cell(row=1, column=1, value="Iteration")
        ws4.cell(row=1, column=2, value="Mean Fitness")
        _style_header(ws4, 1, 2)
        for it in range(len(mean_conv)):
            ws4.cell(row=it+2, column=1, value=it)
            ws4.cell(row=it+2, column=2, value=f"{mean_conv[it]:.10e}")
        _auto_width(ws4)

        # ── Convergence plot ──
        plot_convergence(prob_name, [algo_name],
                         {algo_name: list(mean_conv)},
                         output_dir, max_iter)

        # ── Box plot ──
        plot_boxplot(prob_name, [algo_name],
                     {algo_name: best_fitnesses},
                     output_dir)

    path = os.path.join(output_dir, "Excel Files")
    os.makedirs(path, exist_ok=True)
    wb.save(os.path.join(path, "Engineering.xlsx"))
    print("  Engineering problems complete.")
